import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Constants
EXPORT_FOLDER = os.path.expanduser('~/iccid_merge_project')
MERGED_FILE = os.path.join(EXPORT_FOLDER, 'Merged_Data.xlsx')
FILTERED_MAIN_FILE = os.path.join(EXPORT_FOLDER, 'Filtered_Main_Data.xlsx')
HIGHLIGHTED_MAIN_FILE = os.path.join(EXPORT_FOLDER, 'Main_data_with_highlight.xlsx')
MAIN_DB_FILE = os.path.join(EXPORT_FOLDER, 'Main_data.xlsx')

# Check if export folder exists
if not os.path.exists(EXPORT_FOLDER):
    raise FileNotFoundError(f"Export folder not found: {EXPORT_FOLDER}")

# 1. Merge 18 files and remove duplicates
all_exports = []
for i in range(1, 19):
    file = os.path.join(EXPORT_FOLDER, f'Export_{i}.xlsx')
    if not os.path.exists(file):
        print(f"⚠️ Warning: {file} not found. Skipping.")
        continue
    df = pd.read_excel(file)
    df = df.rename(columns={df.columns[0]: 'ICCID'})  # Standardise column name
    all_exports.append(df[['ICCID']])

if not all_exports:
    raise ValueError("❌ No export files found to merge.")

merged_df = pd.concat(all_exports).drop_duplicates().reset_index(drop=True)
merged_df.to_excel(MERGED_FILE, index=False)
print(f"✅ Step 1: Merged {len(all_exports)} files into {MERGED_FILE}")

# 2. Load main database
if not os.path.exists(MAIN_DB_FILE):
    raise FileNotFoundError(f"Main database file not found: {MAIN_DB_FILE}")

main_df = pd.read_excel(MAIN_DB_FILE)

# 3. Filter main database where ICCID not in merged list
if 'ICCID' not in main_df.columns:
    raise ValueError("'ICCID' column not found in Main_data.xlsx")

filtered_main_df = main_df[~main_df['ICCID'].isin(merged_df['ICCID'])]
filtered_main_df.to_excel(FILTERED_MAIN_FILE, index=False)
print(f"✅ Step 2: Filtered main database written to {FILTERED_MAIN_FILE}")

# 4. Highlight rows in original main database that match merged ICCIDs
wb = load_workbook(MAIN_DB_FILE)
ws = wb.active
highlight_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Dynamically find ICCID column index (1-based for openpyxl)
header = [cell.value for cell in ws[1]]
try:
    iccid_col_idx = header.index('ICCID') + 1
except ValueError:
    raise ValueError("❌ 'ICCID' column not found in header row of Excel file.")

for row in range(2, ws.max_row + 1):
    iccid_value = ws.cell(row=row, column=iccid_col_idx).value
    if iccid_value in merged_df['ICCID'].values:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = highlight_fill

wb.save(HIGHLIGHTED_MAIN_FILE)
print(f"✅ Step 3: Highlighted matching rows saved to {HIGHLIGHTED_MAIN_FILE}")
